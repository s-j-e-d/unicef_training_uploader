import pandas as pd
import re
import hashlib
import os

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo

# ======================
# INPUTS
# ======================
csv_path = "sdp_master_list.csv"
out_path = "sdp_cascade_template.xlsx"

# How many data rows you want available in the form (table will cover this many)
max_rows = 200

# Form header starts at A9 (row 9)
header_row = 9
first_data_row = header_row + 1

# ======================
# Load CSV
# ======================
df = pd.read_csv(csv_path)

# Clean whitespace in user-visible dropdown labels
for col in ["Admin1_Name","Admin2_Name","Admin3_Name","Admin4_Name","parent_phcf_name","sdp_name"]:
    if col in df.columns:
        df[col] = df[col].astype(str).str.strip()

# ======================
# Excel-safe token generator for named ranges
# ======================
def name_token(x: str, maxlen: int = 80) -> str:
    s = str(x).strip()
    s = s.replace("-", "_")
    s = re.sub(r"[^A-Za-z0-9_]", "", s)
    if not s:
        s = "X"
    if len(s) > maxlen:
        h = hashlib.sha1(s.encode("utf-8")).hexdigest()[:8]
        s = s[: maxlen - 9] + "_" + h
    return s

# ======================
# Validate required columns exist
# ======================
required_cols = [
    "Admin1_Pcode","Admin1_Name",
    "Admin2_Pcode","Admin2_Name",
    "Admin3_Pcode","Admin3_Name",
    "Admin4_Pcode","Admin4_Name",
    "parent_phcf_uuid","parent_phcf_name",
    "sdp_uuid","sdp_name"
]
missing = [c for c in required_cols if c not in df.columns]
if missing:
    raise ValueError(f"Missing required columns in CSV: {missing}")

# ======================
# Distinct level tables (sorted alphabetically by display name within parent)
# ======================
admin1 = df[["Admin1_Pcode","Admin1_Name"]].drop_duplicates().copy()
admin1["A1_token"] = admin1["Admin1_Pcode"].map(name_token)
admin1 = admin1.sort_values("Admin1_Name")

admin2 = df[["Admin1_Pcode","Admin2_Pcode","Admin2_Name"]].drop_duplicates().copy()
admin2["A1_token"] = admin2["Admin1_Pcode"].map(name_token)
admin2["A2_token"] = admin2["Admin2_Pcode"].map(name_token)
admin2 = admin2.sort_values(["Admin1_Pcode","Admin2_Name"])

admin3 = df[["Admin2_Pcode","Admin3_Pcode","Admin3_Name"]].drop_duplicates().copy()
admin3["A2_token"] = admin3["Admin2_Pcode"].map(name_token)
admin3["A3_token"] = admin3["Admin3_Pcode"].map(name_token)
admin3 = admin3.sort_values(["Admin2_Pcode","Admin3_Name"])

admin4 = df[["Admin3_Pcode","Admin4_Pcode","Admin4_Name"]].drop_duplicates().copy()
admin4["A3_token"] = admin4["Admin3_Pcode"].map(name_token)
admin4["A4_token"] = admin4["Admin4_Pcode"].map(name_token)
admin4 = admin4.sort_values(["Admin3_Pcode","Admin4_Name"])

phcf = df[["Admin4_Pcode","parent_phcf_uuid","parent_phcf_name"]].drop_duplicates().copy()
phcf["A4_token"] = phcf["Admin4_Pcode"].map(name_token)
phcf["PHCF_token"] = phcf["parent_phcf_uuid"].map(name_token)
phcf = phcf.sort_values(["Admin4_Pcode","parent_phcf_name"])

sdp = df[["parent_phcf_uuid","sdp_uuid","sdp_name"]].drop_duplicates().copy()
sdp["PHCF_token"] = sdp["parent_phcf_uuid"].map(name_token)
sdp = sdp.sort_values(["parent_phcf_uuid","sdp_name"])

# ======================
# Create workbook
# ======================
wb = Workbook()
ws_lists = wb.active
ws_lists.title = "LISTS"
ws_form = wb.create_sheet("FORM")

def write_table(ws, start_row, start_col, headers, rows):
    for j, h in enumerate(headers):
        ws.cell(row=start_row, column=start_col + j, value=h)
    r = start_row + 1
    for row in rows:
        for j, val in enumerate(row):
            ws.cell(row=r, column=start_col + j, value=val)
        r += 1
    return r - 1

# ======================
# LISTS: Mapping tables (parent-aware lookups) + child list blocks
# ======================
r = 1
ws_lists.cell(r, 1, "MAPPING TABLES (do not edit)")
r += 2

# Admin1: Name -> Pcode, Token
end = write_table(
    ws_lists, r, 1,
    ["Admin1_Name","Admin1_Pcode","A1_token"],
    admin1[["Admin1_Name","Admin1_Pcode","A1_token"]].values.tolist()
)
a1_name_rng = (r+1, end, 1, 1)
a1_code_rng = (r+1, end, 2, 2)
a1_tok_rng  = (r+1, end, 3, 3)
r = end + 3

# Admin2: (Admin1_Pcode||Admin2_Name) -> Admin2_Pcode, A2_token
admin2_map = admin2.copy()
admin2_map["k"] = admin2_map["Admin1_Pcode"].astype(str) + "||" + admin2_map["Admin2_Name"].astype(str)
end = write_table(
    ws_lists, r, 1,
    ["Admin1_Pcode||Admin2_Name","Admin2_Pcode","A2_token"],
    admin2_map[["k","Admin2_Pcode","A2_token"]].values.tolist()
)
a2_key_rng  = (r+1, end, 1, 1)
a2_code_rng = (r+1, end, 2, 2)
a2_tok_rng  = (r+1, end, 3, 3)
r = end + 3

# Admin3: (Admin2_Pcode||Admin3_Name) -> Admin3_Pcode, A3_token
admin3_map = admin3.copy()
admin3_map["k"] = admin3_map["Admin2_Pcode"].astype(str) + "||" + admin3_map["Admin3_Name"].astype(str)
end = write_table(
    ws_lists, r, 1,
    ["Admin2_Pcode||Admin3_Name","Admin3_Pcode","A3_token"],
    admin3_map[["k","Admin3_Pcode","A3_token"]].values.tolist()
)
a3_key_rng  = (r+1, end, 1, 1)
a3_code_rng = (r+1, end, 2, 2)
a3_tok_rng  = (r+1, end, 3, 3)
r = end + 3

# Admin4: (Admin3_Pcode||Admin4_Name) -> Admin4_Pcode, A4_token
admin4_map = admin4.copy()
admin4_map["k"] = admin4_map["Admin3_Pcode"].astype(str) + "||" + admin4_map["Admin4_Name"].astype(str)
end = write_table(
    ws_lists, r, 1,
    ["Admin3_Pcode||Admin4_Name","Admin4_Pcode","A4_token"],
    admin4_map[["k","Admin4_Pcode","A4_token"]].values.tolist()
)
a4_key_rng  = (r+1, end, 1, 1)
a4_code_rng = (r+1, end, 2, 2)
a4_tok_rng  = (r+1, end, 3, 3)
r = end + 3

# PHCF: (Admin4_Pcode||PHCF_Name) -> PHCF_UUID, PHCF_token
phcf_map = phcf.copy()
phcf_map["k"] = phcf_map["Admin4_Pcode"].astype(str) + "||" + phcf_map["parent_phcf_name"].astype(str)
end = write_table(
    ws_lists, r, 1,
    ["Admin4_Pcode||PHCF_Name","PHCF_UUID","PHCF_token"],
    phcf_map[["k","parent_phcf_uuid","PHCF_token"]].values.tolist()
)
phcf_key_rng  = (r+1, end, 1, 1)
phcf_uuid_rng = (r+1, end, 2, 2)
phcf_tok_rng  = (r+1, end, 3, 3)
r = end + 3

# SDP: (PHCF_UUID||SDP_Name) -> SDP_UUID
sdp_map = sdp.copy()
sdp_map["k"] = sdp_map["parent_phcf_uuid"].astype(str) + "||" + sdp_map["sdp_name"].astype(str)
end = write_table(
    ws_lists, r, 1,
    ["PHCF_UUID||SDP_Name","SDP_UUID"],
    sdp_map[["k","sdp_uuid"]].values.tolist()
)
sdp_key_rng  = (r+1, end, 1, 1)
sdp_uuid_rng = (r+1, end, 2, 2)
r = end + 4

# Child list blocks
ws_lists.cell(r, 1, "CHILD LIST BLOCKS (named ranges point here)")
r += 2

named_ranges = []

# Root Admin1 list
root_start = r
for i, name in enumerate(admin1["Admin1_Name"].tolist(), start=root_start):
    ws_lists.cell(i, 1, name)
root_end = root_start + len(admin1) - 1
named_ranges.append(("NR_Admin1","LISTS",root_start,1,root_end,1))
r = root_end + 2

# Admin1 -> Admin2: NR_A1_<A1_token>
for a1_pcode, sub in admin2.groupby("Admin1_Pcode"):
    a1_tok = name_token(a1_pcode)
    start = r
    names = sub["Admin2_Name"].drop_duplicates().sort_values().tolist()
    for i, nm in enumerate(names, start=start):
        ws_lists.cell(i, 1, nm)
    end = start + len(names) - 1
    if names:
        named_ranges.append((f"NR_A1_{a1_tok}","LISTS",start,1,end,1))
    r = end + 2

# Admin2 -> Admin3: NR_A2_<A2_token>
for a2_pcode, sub in admin3.groupby("Admin2_Pcode"):
    a2_tok = name_token(a2_pcode)
    start = r
    names = sub["Admin3_Name"].drop_duplicates().sort_values().tolist()
    for i, nm in enumerate(names, start=start):
        ws_lists.cell(i, 1, nm)
    end = start + len(names) - 1
    if names:
        named_ranges.append((f"NR_A2_{a2_tok}","LISTS",start,1,end,1))
    r = end + 2

# Admin3 -> Admin4: NR_A3_<A3_token>
for a3_pcode, sub in admin4.groupby("Admin3_Pcode"):
    a3_tok = name_token(a3_pcode)
    start = r
    names = sub["Admin4_Name"].drop_duplicates().sort_values().tolist()
    for i, nm in enumerate(names, start=start):
        ws_lists.cell(i, 1, nm)
    end = start + len(names) - 1
    if names:
        named_ranges.append((f"NR_A3_{a3_tok}","LISTS",start,1,end,1))
    r = end + 2

# Admin4 -> PHCF: NR_A4_<A4_token>
for a4_pcode, sub in phcf.groupby("Admin4_Pcode"):
    a4_tok = name_token(a4_pcode)
    start = r
    names = sub["parent_phcf_name"].drop_duplicates().sort_values().tolist()
    for i, nm in enumerate(names, start=start):
        ws_lists.cell(i, 1, nm)
    end = start + len(names) - 1
    if names:
        named_ranges.append((f"NR_A4_{a4_tok}","LISTS",start,1,end,1))
    r = end + 2

# PHCF -> SDP: NR_PHCF_<PHCF_token>
for phcf_uuid, sub in sdp.groupby("parent_phcf_uuid"):
    phcf_tok = name_token(phcf_uuid)
    start = r
    names = sub["sdp_name"].drop_duplicates().sort_values().tolist()
    for i, nm in enumerate(names, start=start):
        ws_lists.cell(i, 1, nm)
    end = start + len(names) - 1
    if names:
        named_ranges.append((f"NR_PHCF_{phcf_tok}","LISTS",start,1,end,1))
    r = end + 1

# Create defined names
for nm, sheet, r1, c1, r2, c2 in named_ranges:
    addr = f"'{sheet}'!${get_column_letter(c1)}${r1}:${get_column_letter(c2)}${r2}"
    wb.defined_names.add(DefinedName(nm, attr_text=addr))

# ======================
# FORM: Your exact headers starting at A9
# ======================
headers = [
    "First Name",
    "Second Name",
    "Phone Number",
    "Email",
    "Gender",
    "Are you working in Health Facility (Yes or No) / \nМісце роботи - Заклад Охорони Здоров'я (Так чи ні)",
    "Position / \nПосада",
    "Oblast of Parent PHCF",
    "Rayon of Parent PHCF",
    "Hromada of Parent PHCF",
    "Settlement of Parent PHCF",
    "Name of PHCF",
    "Name of SDP",
    "Other Service Provider or Facility / Інше Місце надання послуг чи Заклад",
    "Post-Training Test Score %",
    "Training start date\nmm/yyyy",
    "Training end date\nmm/yyyy",
    "Name of Training",
    "Place where the training is conducted"
]

# Write headers at row 9
for j, h in enumerate(headers, start=1):
    ws_form.cell(row=header_row, column=j, value=h)

# Helpful UX
ws_form.freeze_panes = f"A{first_data_row}"

# Column widths (adjust as you like)
for col_letter, width in {
    "A":14, "B":16, "C":16, "D":26, "E":10, "F":34, "G":18,
    "H":22, "I":22, "J":22, "K":24, "L":34, "M":36,
    "N":28, "O":22, "P":18, "Q":18, "R":26, "S":26
}.items():
    ws_form.column_dimensions[col_letter].width = width

# ======================
# Hidden helper columns to the right (tokens/IDs)
# We'll place them starting at column T (20)
# ======================
helper_headers = [
    "Admin1_Pcode", "Admin2_Pcode", "Admin3_Pcode", "Admin4_Pcode", "PHCF_UUID", "SDP_UUID",
    "A1_token", "A2_token", "A3_token", "A4_token", "PHCF_token"
]
helper_start_col = 20  # T
for j, h in enumerate(helper_headers, start=helper_start_col):
    ws_form.cell(row=header_row, column=j, value=h)

# Hide helper columns T:AD
for col in range(helper_start_col, helper_start_col + len(helper_headers)):
    ws_form.column_dimensions[get_column_letter(col)].hidden = True

# ======================
# FORM formulas for helper columns (row-by-row)
# Cascade visible columns:
#   H Admin1_Name (Oblast)
#   I Admin2_Name (Rayon)
#   J Admin3_Name (Hromada)
#   K Admin4_Name (Settlement)
#   L PHCF_Name
#   M SDP_Name
# ======================
# Helper columns indices
col_Admin1_Pcode = helper_start_col + 0  # T
col_Admin2_Pcode = helper_start_col + 1  # U
col_Admin3_Pcode = helper_start_col + 2  # V
col_Admin4_Pcode = helper_start_col + 3  # W
col_PHCF_UUID    = helper_start_col + 4  # X
col_SDP_UUID     = helper_start_col + 5  # Y
col_A1_token     = helper_start_col + 6  # Z
col_A2_token     = helper_start_col + 7  # AA
col_A3_token     = helper_start_col + 8  # AB
col_A4_token     = helper_start_col + 9  # AC
col_PHCF_token   = helper_start_col + 10 # AD

# Row range
last_data_row = header_row + max_rows

for rr in range(first_data_row, last_data_row + 1):
    # Admin1_Pcode + A1_token from Admin1_Name (H)
    ws_form.cell(rr, col_Admin1_Pcode, value=
        f'=IFERROR(INDEX(LISTS!$B${a1_code_rng[0]}:$B${a1_code_rng[1]}, '
        f'MATCH($H{rr}, LISTS!$A${a1_name_rng[0]}:$A${a1_name_rng[1]}, 0)),"")'
    )
    ws_form.cell(rr, col_A1_token, value=
        f'=IFERROR(INDEX(LISTS!$C${a1_tok_rng[0]}:$C${a1_tok_rng[1]}, '
        f'MATCH($H{rr}, LISTS!$A${a1_name_rng[0]}:$A${a1_name_rng[1]}, 0)),"")'
    )

    # Admin2_Pcode + A2_token from (Admin1_Pcode || Admin2_Name) where Admin2_Name is I
    ws_form.cell(rr, col_Admin2_Pcode, value=
        f'=IFERROR(INDEX(LISTS!$B${a2_code_rng[0]}:$B${a2_code_rng[1]}, '
        f'MATCH(${get_column_letter(col_Admin1_Pcode)}{rr}&"||"&$I{rr}, LISTS!$A${a2_key_rng[0]}:$A${a2_key_rng[1]}, 0)),"")'
    )
    ws_form.cell(rr, col_A2_token, value=
        f'=IFERROR(INDEX(LISTS!$C${a2_tok_rng[0]}:$C${a2_tok_rng[1]}, '
        f'MATCH(${get_column_letter(col_Admin1_Pcode)}{rr}&"||"&$I{rr}, LISTS!$A${a2_key_rng[0]}:$A${a2_key_rng[1]}, 0)),"")'
    )

    # Admin3_Pcode + A3_token from (Admin2_Pcode || Admin3_Name) where Admin3_Name is J
    ws_form.cell(rr, col_Admin3_Pcode, value=
        f'=IFERROR(INDEX(LISTS!$B${a3_code_rng[0]}:$B${a3_code_rng[1]}, '
        f'MATCH(${get_column_letter(col_Admin2_Pcode)}{rr}&"||"&$J{rr}, LISTS!$A${a3_key_rng[0]}:$A${a3_key_rng[1]}, 0)),"")'
    )
    ws_form.cell(rr, col_A3_token, value=
        f'=IFERROR(INDEX(LISTS!$C${a3_tok_rng[0]}:$C${a3_tok_rng[1]}, '
        f'MATCH(${get_column_letter(col_Admin2_Pcode)}{rr}&"||"&$J{rr}, LISTS!$A${a3_key_rng[0]}:$A${a3_key_rng[1]}, 0)),"")'
    )

    # Admin4_Pcode + A4_token from (Admin3_Pcode || Admin4_Name) where Admin4_Name is K
    ws_form.cell(rr, col_Admin4_Pcode, value=
        f'=IFERROR(INDEX(LISTS!$B${a4_code_rng[0]}:$B${a4_code_rng[1]}, '
        f'MATCH(${get_column_letter(col_Admin3_Pcode)}{rr}&"||"&$K{rr}, LISTS!$A${a4_key_rng[0]}:$A${a4_key_rng[1]}, 0)),"")'
    )
    ws_form.cell(rr, col_A4_token, value=
        f'=IFERROR(INDEX(LISTS!$C${a4_tok_rng[0]}:$C${a4_tok_rng[1]}, '
        f'MATCH(${get_column_letter(col_Admin3_Pcode)}{rr}&"||"&$K{rr}, LISTS!$A${a4_key_rng[0]}:$A${a4_key_rng[1]}, 0)),"")'
    )

    # PHCF_UUID + PHCF_token from (Admin4_Pcode || PHCF_Name) where PHCF_Name is L
    ws_form.cell(rr, col_PHCF_UUID, value=
        f'=IFERROR(INDEX(LISTS!$B${phcf_uuid_rng[0]}:$B${phcf_uuid_rng[1]}, '
        f'MATCH(${get_column_letter(col_Admin4_Pcode)}{rr}&"||"&$L{rr}, LISTS!$A${phcf_key_rng[0]}:$A${phcf_key_rng[1]}, 0)),"")'
    )
    ws_form.cell(rr, col_PHCF_token, value=
        f'=IFERROR(INDEX(LISTS!$C${phcf_tok_rng[0]}:$C${phcf_tok_rng[1]}, '
        f'MATCH(${get_column_letter(col_Admin4_Pcode)}{rr}&"||"&$L{rr}, LISTS!$A${phcf_key_rng[0]}:$A${phcf_key_rng[1]}, 0)),"")'
    )

    # SDP_UUID from (PHCF_UUID || SDP_Name) where SDP_Name is M
    ws_form.cell(rr, col_SDP_UUID, value=
        f'=IFERROR(INDEX(LISTS!$B${sdp_uuid_rng[0]}:$B${sdp_uuid_rng[1]}, '
        f'MATCH(${get_column_letter(col_PHCF_UUID)}{rr}&"||"&$M{rr}, LISTS!$A${sdp_key_rng[0]}:$A${sdp_key_rng[1]}, 0)),"")'
    )

# ======================
# Data validation for cascade columns H..M applied to ALL rows
# IMPORTANT: formulas use the first data row index, but are applied across the full range.
# ======================
r0 = first_data_row  # base row for DV formula row-relative reference

dv_admin1 = DataValidation(type="list", formula1="=NR_Admin1", allow_blank=True)
dv_admin2 = DataValidation(type="list",
                           formula1=f'=IF(${get_column_letter(col_A1_token)}{r0}="","",INDIRECT("NR_A1_"&${get_column_letter(col_A1_token)}{r0}))',
                           allow_blank=True)
dv_admin3 = DataValidation(type="list",
                           formula1=f'=IF(${get_column_letter(col_A2_token)}{r0}="","",INDIRECT("NR_A2_"&${get_column_letter(col_A2_token)}{r0}))',
                           allow_blank=True)
dv_admin4 = DataValidation(type="list",
                           formula1=f'=IF(${get_column_letter(col_A3_token)}{r0}="","",INDIRECT("NR_A3_"&${get_column_letter(col_A3_token)}{r0}))',
                           allow_blank=True)
dv_phcf   = DataValidation(type="list",
                           formula1=f'=IF(${get_column_letter(col_A4_token)}{r0}="","",INDIRECT("NR_A4_"&${get_column_letter(col_A4_token)}{r0}))',
                           allow_blank=True)
dv_sdp    = DataValidation(type="list",
                           formula1=f'=IF(${get_column_letter(col_PHCF_token)}{r0}="","",INDIRECT("NR_PHCF_"&${get_column_letter(col_PHCF_token)}{r0}))',
                           allow_blank=True)

for dv in [dv_admin1, dv_admin2, dv_admin3, dv_admin4, dv_phcf, dv_sdp]:
    ws_form.add_data_validation(dv)

# Apply to full ranges (H..M)
dv_admin1.add(f"H{first_data_row}:H{last_data_row}")
dv_admin2.add(f"I{first_data_row}:I{last_data_row}")
dv_admin3.add(f"J{first_data_row}:J{last_data_row}")
dv_admin4.add(f"K{first_data_row}:K{last_data_row}")
dv_phcf.add(f"L{first_data_row}:L{last_data_row}")
dv_sdp.add(f"M{first_data_row}:M{last_data_row}")

# ======================
# Make visible area (A9:S...) an Excel Table
# ======================
last_visible_col = 19  # A..S
table_ref = f"A{header_row}:{get_column_letter(last_visible_col)}{last_data_row}"
tbl = Table(displayName="tFORM", ref=table_ref)
tbl.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)
ws_form.add_table(tbl)

# Hide LISTS sheet
ws_lists.sheet_state = "hidden"

# Save (remove existing file first to avoid OneDrive/Excel lock confusion)
# Note: will still fail if file is open in Excel.
if os.path.exists(out_path):
    os.remove(out_path)

wb.save(out_path)
print(f"Wrote: {out_path}")
