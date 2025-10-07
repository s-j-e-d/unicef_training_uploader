from __future__ import annotations

import json
import re
import uuid
import os
import io
from typing import Any, Dict, List, Optional

import pandas as pd
import streamlit as st
import requests
import warnings
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

warnings.filterwarnings("ignore", message="Data Validation extension is not supported and will be removed")

# -------------------------------
# CONFIG HELPERS
# -------------------------------
def get_secret(name: str, default: str = "") -> str:
    # Lookup order: env -> st.secrets -> default
    return os.environ.get(name) or st.secrets.get(name, default)

# Defaults for EU
EU_ASSET_UID = "aS5UNu8wuynhJN87JYopFQ"
EU_SUBMIT_URL = get_secret("EU_SUBMIT_URL", "https://kc-eu.kobotoolbox.org/api/v1/submissions.json")
# NOTE: we no longer hard-stop on KOBO_TOKEN at import time; we check after the user picks a form.

FORMS: Dict[str, Dict[str, Any]] = {
    "Training Attendance (EU)": {
        "asset_uid": EU_ASSET_UID,
        "submit_url": EU_SUBMIT_URL,
        "api_key_env": "KOBO_TOKEN",
        "landing": "https://eu.kobotoolbox.org/#/forms/aS5UNu8wuynhJN87JYopFQ/landing",
        "auth_scheme": "Token",
    },
    "Training Attendance (Ruslan)": {
        # Set these in your environment/secrets
        "asset_uid": get_secret("RUSLAN_ASSET_UID", ""),               # REQUIRED
        "submit_url": get_secret("RUSLAN_SUBMIT_URL", ""),             # REQUIRED
        "api_key_env": "KOBO_TOKEN_RUSLAN",                            # REQUIRED
        "landing": get_secret("RUSLAN_LANDING", ""),                   # optional
        "auth_scheme": "Token",                                        # or "Bearer" if needed
    },
}

# -------------------------------
# EXACT column headers from TblAttend
# -------------------------------
H = {
    "first_name": "First Name",
    "second_name": "Second Name",
    "phone": "Phone Number",
    "email": "Email",
    "gender": "Gender",
    "works_hf": "Are you working in Health Facility (Yes or No) / \nМісце роботи - Заклад Охорони Здоров'я (Так чи ні)",
    "position": "Position / \nПосада",

    "oblast_parent": "Oblast of Parent PCHF",
    "rayon_parent": "Rayon of Parent PCHF",
    "hromada_parent": "Hromada of Parent PCHF",
    "parent_pchf_name": "Name of Parent PCHF",
    "ambulatory_name": "Name of Ambulatory",

    "helper": "Helper",
    "service_provider_other": "Other Service Provider or Facility / Інше Місце надання послуг чи Заклад",

    "score": "Post-Training Test Score %",
    "start_mm_yyyy": "Training start date\n mm/yyyy",
    "end_mm_yyyy": "Training end date\n mm/yyyy",
    "training_name": "Name of Training",
    "place": "Place where the training is conducted",
}

# (K is retained if you use it elsewhere)
K = {
    "first_name": "personal_info/first_name",
    "second_name": "personal_info/second_name",
    "phone": "personal_info/phone",
    "email": "personal_info/email",
    "gender": "personal_info/gender",
    "works_hf": "personal_info/works_in_hf",

    "position": "work_location/position",
    "oblast": "work_location/oblast",
    "rayon": "work_location/rayon",
    "hromada": "work_location/hromada",
    "settlement": "work_location/settlement",
    "facility_code": "work_location/facility_code",
    "service_provider": "work_location/service_provider",
    "helper": "work_location/helper",
    "service_provider_other": "work_location/service_provider_other",

    "score": "training_score/training_score_pct",
    "start_date": "metadata/start_date",
    "end_date": "metadata/end_date",
    "training_name": "metadata/training_name",
    "place": "metadata/place",
}

# -------------------------------
# Helpers
# -------------------------------
def normalize_yes_no(v: Any) -> Optional[str]:
    if v is None or (isinstance(v, float) and pd.isna(v)): return None
    s = str(v).strip().lower()
    if s in {"yes", "y", "true", "1", "так"}: return "yes"
    if s in {"no", "n", "false", "0", "ні"}: return "no"
    return None

def normalize_gender(v: Any) -> Optional[str]:
    if v is None or (isinstance(v, float) and pd.isna(v)): return None
    s = str(v).strip().lower()
    if s.startswith("f"): return "female"
    if s.startswith("m"): return "male"
    if "other" in s: return "other"
    if "prefer" in s: return "prefer_not_to_say"
    return None

def parse_mm_yyyy_or_date(v: Any) -> Optional[str]:
    if v is None or (isinstance(v, float) and pd.isna(v)): return None
    ts = pd.to_datetime(v, errors="coerce", dayfirst=True)
    if pd.notna(ts):
        return ts.strftime("%Y-%m-%d")
    s = str(v).strip()
    m2 = re.match(r"^(\d{1,2})\s*[-/.]\s*(\d{4})$", s)  # mm/yyyy
    if m2:
        mm, yyyy = int(m2.group(1)), int(m2.group(2))
        try:
            return pd.Timestamp(year=yyyy, month=mm, day=1).strftime("%Y-%m-%d")
        except Exception:
            return None
    return None

def parse_score_int(v: Any) -> Optional[int]:
    if v is None or (isinstance(v, float) and pd.isna(v)): return None
    s = str(v).strip().replace("%", "")
    try:
        f = float(s)
    except Exception:
        return None
    n = int(round(f))
    return n if 0 <= n <= 100 else None

def prune(d: Dict[str, Any]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for k, v in d.items():
        if isinstance(v, dict):
            pv = prune(v)
            if pv: out[k] = pv
        else:
            if v is None: continue
            if isinstance(v, float) and pd.isna(v): continue
            if isinstance(v, str) and v.strip() == "": continue
            out[k] = v
    return out

def post_submission_json(submit_url: str, api_key: str, auth_scheme: str, asset_uid: str, submission: Dict[str, Any]) -> requests.Response:
    if not api_key:
        raise RuntimeError("Missing API key for the selected form.")
    headers = {
        "Authorization": f"{auth_scheme} {api_key}",
        "Accept": "application/json",
        "Content-Type": "application/json",
    }
    body = {"id": asset_uid, "submission": submission}
    return requests.post(submit_url, headers=headers, data=json.dumps(body), timeout=60)

# -------------------------------
# Load TblAttend
# -------------------------------
def load_tbl_attend(file_like, sheet="Attend", table="TblAttend") -> pd.DataFrame:
    raw = file_like.read()
    wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=False)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet}' not found")
    ws = wb[sheet]
    if table not in ws.tables:
        raise ValueError(f"Table '{table}' not found in sheet '{sheet}'")
    ref = ws.tables[table].ref
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    rows = list(ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True))
    if not rows:
        raise ValueError("TblAttend is empty")
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data = rows[1:]
    df = pd.DataFrame(data, columns=headers).dropna(how="all")
    return df

# -------------------------------
# Build one submission (nested exactly as schema)
# -------------------------------
def build_submission(row: pd.Series) -> Dict[str, Any]:
    sub = {
        "metadata": {
            "training_name": None if pd.isna(row.get(H["training_name"])) else str(row.get(H["training_name"])).strip(),
            "start_date": parse_mm_yyyy_or_date(row.get(H["start_mm_yyyy"])),
            "end_date": parse_mm_yyyy_or_date(row.get(H["end_mm_yyyy"])),
            "place": None if pd.isna(row.get(H["place"])) else str(row.get(H["place"])).strip(),
        },
        "personal_info": {
            "first_name": None if pd.isna(row.get(H["first_name"])) else str(row.get(H["first_name"])).strip(),
            "second_name": None if pd.isna(row.get(H["second_name"])) else str(row.get(H["second_name"])).strip(),
            "phone": None if pd.isna(row.get(H["phone"])) else str(row.get(H["phone"])).strip(),
            "email": None if pd.isna(row.get(H["email"])) else str(row.get(H["email"])).strip(),
            "gender": normalize_gender(row.get(H["gender"])),
            "works_in_hf": normalize_yes_no(row.get(H["works_hf"])),
        },
        "work_location": {
            "position": None if pd.isna(row.get(H["position"])) else str(row.get(H["position"])).strip(),
            "oblast": None if pd.isna(row.get(H["oblast_parent"])) else str(row.get(H["oblast_parent"])).strip(),
            "rayon": None if pd.isna(row.get(H["rayon_parent"])) else str(row.get(H["rayon_parent"])).strip(),
            "hromada": None if pd.isna(row.get(H["hromada_parent"])) else str(row.get(H["hromada_parent"])).strip(),
            "settlement": None,
            "facility_code": None if pd.isna(row.get(H["parent_pchf_name"])) else str(row.get(H["parent_pchf_name"])).strip(),
            "service_provider": None if pd.isna(row.get(H["ambulatory_name"])) else str(row.get(H["ambulatory_name"])).strip(),
            "helper": None if pd.isna(row.get(H["helper"])) else str(row.get(H["helper"])).strip(),
            "service_provider_other": None if pd.isna(row.get(H["service_provider_other"])) else str(row.get(H["service_provider_other"])).strip(),
        },
        "training_score": {
            "training_score_pct": parse_score_int(row.get(H["score"])),
        },
        "meta": {"instanceID": f"uuid:{uuid.uuid4()}"},
    }
    return prune(sub)

# -------------------------------
# Minimal Validation (hard-coded)
# -------------------------------
def is_blank(v: Any) -> bool:
    if v is None: return True
    if isinstance(v, float) and pd.isna(v): return True
    s = str(v).strip()
    return s == "" or s.lower() == "nan" or s == "NaT"

REQUIRED_COLS = [
    H["first_name"], H["second_name"], H["phone"], H["email"], H["gender"], H["works_hf"],
    H["position"], H["oblast_parent"], H["rayon_parent"], H["hromada_parent"],
    H["score"], H["start_mm_yyyy"], H["end_mm_yyyy"], H["training_name"], H["place"]
]

def validate(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    issues: List[Dict[str, Any]] = []
    for i, row in df.iterrows():
        errs: List[str] = []
        for c in REQUIRED_COLS:
            if c not in df.columns or is_blank(row.get(c)):
                errs.append(f"Missing: {c}")
        if (H["parent_pchf_name"] in df.columns and is_blank(row.get(H["parent_pchf_name"]))) and \
           (H["ambulatory_name"] in df.columns and is_blank(row.get(H["ambulatory_name"]))):
            errs.append(f"Provide at least one of '{H['parent_pchf_name']}' or '{H['ambulatory_name']}'")
        sv = row.get(H["score"]) if H["score"] in df.columns else None
        if parse_score_int(sv) is None:
            errs.append("Score must be a whole number 0–100")
        for c in (H["start_mm_yyyy"], H["end_mm_yyyy"]):
            if not is_blank(row.get(c)) and parse_mm_yyyy_or_date(row.get(c)) is None:
                errs.append(f"Invalid date in: {c} (use Excel date or mm/yyyy or dd-mm-yyyy)")
        if errs:
            issues.append({"Row #": i+1, "Name": row.get(H["first_name"]), "Errors": "; ".join(errs)})
    iss_df = pd.DataFrame(issues) if issues else pd.DataFrame(columns=["Row #","Name","Errors"])
    return iss_df, len(issues)

# -------------------------------
# UI
# -------------------------------
st.set_page_config(page_title="Kobo Uploader", page_icon="⬆️", layout="centered")
st.title("⬆️ Excel → Kobo submission")

form_choice = st.selectbox("Destination form", list(FORMS.keys()))
cfg = FORMS[form_choice]

asset_uid = cfg.get("asset_uid", "").strip()
submit_url = (cfg.get("submit_url") or "").strip()
api_key_env = cfg.get("api_key_env", "KOBO_TOKEN").strip()
auth_scheme = (cfg.get("auth_scheme") or "Token").strip()
api_key = get_secret(api_key_env, "").strip()

if cfg.get("landing"):
    st.caption(f"Form: {cfg['landing']}")
st.caption(f"Submit URL: {submit_url or '—'}")
st.caption(f"API key source: {api_key_env} ({'found' if api_key else 'missing'})")

# Guardrails per selected form
if not asset_uid:
    st.error("Missing asset UID for the selected form. Set RUSLAN_ASSET_UID (for Ruslan) or check config.")
    st.stop()
if not submit_url:
    st.error("Missing submit URL for the selected form. Set RUSLAN_SUBMIT_URL (for Ruslan) or check config.")
    st.stop()
if not api_key:
    st.error(f"Missing API key for the selected form. Set '{api_key_env}' in your environment or Streamlit secrets.")
    st.stop()

uploaded = st.file_uploader("Upload UNICEF Attendance Sheet (.xlsx)", type=["xlsx"])

if uploaded:
    try:
        df = load_tbl_attend(uploaded, sheet="Attend", table="TblAttend")
    except Exception as e:
        st.error(str(e))
        st.stop()

    st.subheader("Preview (first 10 rows)")
    st.dataframe(df.head(10), use_container_width=True)

    val_df, n_err = validate(df)
    if n_err:
        st.error(f"{n_err} row(s) have errors. Fix and re-upload.")
        if not val_df.empty:
            st.dataframe(val_df, use_container_width=True)
        st.stop()
    else:
        st.success("Validation passed.")

    submissions = [build_submission(row) for _, row in df.iterrows()]
    st.write(f"Prepared {len(submissions)} submission(s)")

    if st.button("🚀 Submit to Kobo", type="primary", use_container_width=True):
        progress = st.progress(0)
        status = st.empty()
        succ, fail = 0, 0
        results: List[Dict[str, Any]] = []

        for i, sub in enumerate(submissions, start=1):
            try:
                r = post_submission_json(submit_url, api_key, auth_scheme, asset_uid, sub)
                ok = r.status_code in (200, 201, 202)
                data = r.json() if "application/json" in (r.headers.get("content-type","")) else {"status": r.status_code, "text": r.text[:800]}
            except Exception as e:
                ok = False
                data = {"error": str(e)}
                r = None

            results.append({"ok": ok, "status": getattr(r, "status_code", "n/a"), "response": data})
            succ += 1 if ok else 0
            fail += 0 if ok else 1
            progress.progress(i/len(submissions))
            status.write(f"Uploaded {i}/{len(submissions)} … (ok: {succ}, failed: {fail})")

        st.write({"successes": succ, "failures": fail})
        st.json(results[:5])
        if fail == 0:
            st.success("All submissions sent successfully.")
        else:
            st.warning("Some submissions failed. See responses above.")
else:
    st.info("Upload an Excel file to begin.")
